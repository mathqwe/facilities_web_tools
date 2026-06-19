from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.rateia_ai.core.escritor_excel import criar_workbook_rateio
from src.rateia_ai.core.modelo import ErroRateio, LinhaRateio, RateioArquivo
from src.rateia_ai.core.normalizacao import (
    decimal_moeda,
    descricao_por_cc_ou_texto,
    extrair_cc_de_texto,
    formatar_moeda_br,
    identificador_de_nome_arquivo,
    limpar_identificador,
    nome_seguro_rateio,
    normalizar_texto,
    para_decimal,
)
from src.rateia_ai.core.rateador import somar_por_cc

ROTULO_SEM_CC = "SEM CC"


def _pontuar_headers(headers: list[Any]) -> int:
    norm = [normalizar_texto(h) for h in headers]
    score = 0
    if any("liquido" in h and "cliente" in h for h in norm) or any("faturado" in h for h in norm):
        score += 3
    if any("custo" in h and "cliente" in h for h in norm):
        score += 3
    if any("ft" in h for h in norm):
        score += 1
    return score


def _localizar_cabecalho(ws) -> tuple[int, list[Any]]:
    melhor = (0, -1, [])
    for r in range(1, min(ws.max_row, 30) + 1):
        headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        score = _pontuar_headers(headers)
        if score > melhor[1]:
            melhor = (r, score, headers)
    if melhor[1] < 5:
        raise ErroRateio("Não consegui localizar o cabeçalho da planilha Biztrip/Toptur.")
    return melhor[0], melhor[2]


def _colunas(headers: list[Any], predicado) -> list[int]:
    return [i for i, h in enumerate(headers, start=1) if predicado(normalizar_texto(h))]


def _coluna_valor(headers: list[Any]) -> int | None:
    candidatos = _colunas(headers, lambda h: ("liquido" in h and "cliente" in h) or ("faturado" in h and ("cli" in h or "cliente" in h)) or h == "valor")
    return candidatos[0] if candidatos else None


def _colunas_cc(headers: list[Any]) -> list[int]:
    candidatos = _colunas(headers, lambda h: ("custo" in h and "cliente" in h) or ("centro" in h and "custo" in h) or h in {"projeto", "obra"})
    def prioridade(col: int) -> int:
        h = normalizar_texto(headers[col - 1])
        if "desc" in h or "descricao" in h: return 0
        if "cod" in h or "codigo" in h: return 1
        return 2
    return sorted(dict.fromkeys(candidatos), key=prioridade)


def _coluna_ft(headers: list[Any]) -> int | None:
    candidatos = _colunas(headers, lambda h: "ft" in h or ("nrs" in h and "cli" in h))
    return candidatos[0] if candidatos else None


def _escolher_aba(wb):
    melhor_ws = None
    melhor_score = -1
    melhor_headers = []
    melhor_header_row = 1
    for ws in wb.worksheets:
        if normalizar_texto(ws.title) == "rateio":
            continue
        try:
            row, headers = _localizar_cabecalho(ws)
            score = _pontuar_headers(headers)
        except Exception:
            continue
        if score > melhor_score:
            melhor_ws = ws
            melhor_score = score
            melhor_headers = headers
            melhor_header_row = row
    if not melhor_ws:
        raise ErroRateio("Não encontrei uma aba de base válida para Biztrip/Toptur.")
    return melhor_ws, melhor_header_row, melhor_headers


def _filtrar_por_ft(ws, header_row: int, col_ft: int | None, numero: str) -> set[int] | None:
    if not col_ft or numero == "SEM_ID":
        return None
    linhas = set()
    for r in range(header_row + 1, ws.max_row + 1):
        num = limpar_identificador(ws.cell(r, col_ft).value, 3)
        if num == numero:
            linhas.add(r)
    return linhas or None


def processar(caminho_entrada: Path, pasta_saida: Path, overwrite: bool = True) -> list[RateioArquivo]:
    caminho_entrada = Path(caminho_entrada)
    if caminho_entrada.suffix.lower() != ".xlsx":
        raise ErroRateio("Biztrip/Toptur aceita apenas arquivos .xlsx.")

    wb = load_workbook(caminho_entrada, data_only=True)
    try:
        ws, header_row, headers = _escolher_aba(wb)
        col_valor = _coluna_valor(headers)
        origem_cc_cols = _colunas_cc(headers)
        col_ft = _coluna_ft(headers)
        if not col_valor:
            raise ErroRateio("Não encontrei a coluna de valor da Biztrip/Toptur.")
        if not origem_cc_cols:
            raise ErroRateio("Não encontrei a coluna de centro de custo da Biztrip/Toptur.")

        identificador = nome_seguro_rateio(identificador_de_nome_arquivo(caminho_entrada))
        linhas_permitidas = _filtrar_por_ft(ws, header_row, col_ft, identificador)

        headers_base = [str(h or "") for h in headers] + ["CC AJUSTADO", "CC DESCRIÇÃO"]
        dados_base: list[list[Any]] = []
        linhas: list[LinhaRateio] = []
        linhas_sem_cc = 0
        valor_sem_cc = Decimal("0.00")
        linhas_ignoradas = 0
        total_arquivo = Decimal("0.00")

        for r in range(header_row + 1, ws.max_row + 1):
            if linhas_permitidas is not None and r not in linhas_permitidas:
                linhas_ignoradas += 1
                continue
            valores = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
            valor = para_decimal(ws.cell(r, col_valor).value)
            fontes = [ws.cell(r, c).value for c in origem_cc_cols]
            if all(v in (None, "") for v in valores) or (valor == Decimal("0.00") and all(f in (None, "") for f in fontes)):
                continue
            cc = None
            fonte_usada = None
            for fonte in fontes:
                cc = extrair_cc_de_texto(fonte)
                if cc:
                    fonte_usada = fonte
                    break
            descricao = descricao_por_cc_ou_texto(cc, fonte_usada or (fontes[0] if fontes else None))
            if not cc and valor != Decimal("0.00"):
                cc_rateio = ROTULO_SEM_CC
                linhas_sem_cc += 1
                valor_sem_cc += decimal_moeda(valor)
            else:
                cc_rateio = cc
            incluir = cc_rateio is not None and valor != Decimal("0.00")
            linhas.append(LinhaRateio(fontes, cc_rateio, descricao, valor, incluir_no_rateio=incluir))
            total_arquivo += decimal_moeda(valor)
            dados_base.append(valores + [int(cc_rateio) if str(cc_rateio).isdigit() else cc_rateio, descricao])
    finally:
        wb.close()

    nome_saida = f"RATEIO_{identificador}.xlsx"
    caminho_saida = pasta_saida / nome_saida
    if caminho_saida.exists() and not overwrite:
        base = caminho_saida.with_suffix("")
        contador = 2
        while caminho_saida.exists():
            caminho_saida = base.with_name(f"{base.name}_{contador}").with_suffix(".xlsx")
            contador += 1
        nome_saida = caminho_saida.name

    totais = somar_por_cc(linhas)
    criar_workbook_rateio(caminho_saida, f"RATEIO {identificador}", "BASE", headers_base, dados_base, totais)
    avisos = []
    if linhas_sem_cc:
        avisos.append(f"{linhas_sem_cc} linha(s) sem CC foram agrupadas como SEM CC ({formatar_moeda_br(valor_sem_cc)}).")
    if linhas_ignoradas and linhas_permitidas is not None:
        avisos.append(f"{linhas_ignoradas} linha(s) foram ignoradas por pertencerem a outra FT/fatura.")
    total_rateado = sum(totais.values(), Decimal("0.00"))
    return [
        RateioArquivo(
            fornecedor="Biztrip/Toptur",
            subtipo="FT/FATURA",
            arquivo_origem=caminho_entrada.name,
            identificador=identificador,
            nome_saida=nome_saida,
            arquivo_saida=caminho_saida,
            total_rateado=decimal_moeda(total_rateado),
            total_arquivo=decimal_moeda(total_arquivo),
            valor_sem_cc=decimal_moeda(valor_sem_cc),
            linhas_processadas=len(linhas),
            linhas_rateadas=sum(1 for l in linhas if l.incluir_no_rateio and l.cc_ajustado),
            linhas_sem_cc=linhas_sem_cc,
            quantidade_cc=len(totais),
            avisos=avisos,
        )
    ]
