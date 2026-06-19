from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.rateia_ai.core.escritor_excel import criar_workbook_rateio
from src.rateia_ai.core.modelo import ErroRateio, LinhaRateio, RateioArquivo
from src.rateia_ai.core.normalizacao import (
    decimal_moeda,
    descricao_por_cc_ou_texto,
    extrair_cc_de_texto,
    formatar_moeda_br,
    nome_seguro_rateio,
    normalizar_texto,
    para_decimal,
)
from src.rateia_ai.core.rateador import somar_por_cc


CONFIG_ABAS = {
    "LOCACAO": {
        "nomes": ["locacao", "locação"],
        "coluna_valor": ["total fatura"],
        "coluna_cc": ["centro de custo"],
        "coluna_fatura": ["fatura"],
        "coluna_nome": ["nome do condutor"],
        "subtipo": "LOCAÇÃO",
    },
    "MULTAS": {
        "nomes": ["multas", "multa"],
        "coluna_valor": ["valor da multa"],
        "coluna_cc": ["centro de custo"],
        "coluna_fatura": ["fatura"],
        "coluna_nome": ["cliente"],
        "subtipo": "MULTAS",
    },
}


def _headers(ws, header_row: int = 1) -> list[Any]:
    return [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]


def _coluna(headers: list[Any], nomes: list[str]) -> int | None:
    nomes_norm = [normalizar_texto(n) for n in nomes]
    for idx, header in enumerate(headers, start=1):
        h = normalizar_texto(header)
        if h in nomes_norm:
            return idx
    for idx, header in enumerate(headers, start=1):
        h = normalizar_texto(header)
        if any(nome in h for nome in nomes_norm):
            return idx
    return None


def _aba_por_tipo(wb, tipo: str):
    config = CONFIG_ABAS[tipo]
    for ws in wb.worksheets:
        nome = normalizar_texto(ws.title)
        if any(normalizar_texto(n) == nome for n in config["nomes"]):
            return ws
    for ws in wb.worksheets:
        nome = normalizar_texto(ws.title)
        if any(normalizar_texto(n) in nome for n in config["nomes"]):
            return ws
    return None


def _linha_total_ou_vazia(valores: list[Any], idx_valor: int) -> bool:
    """Remove linhas de total geral da medição.

    Nos arquivos Movida enviados, há uma linha no fim com apenas o total da
    fatura. Ela não pode entrar no rateio, senão o total dobra.
    """
    valor = valores[idx_valor - 1] if idx_valor <= len(valores) else None
    if valor in (None, ""):
        return True
    outros = [v for i, v in enumerate(valores, start=1) if i != idx_valor and v not in (None, "")]
    return len(outros) == 0




def _deve_ocultar_zero_na_base(header: Any, col_idx: int, col_valor: int) -> bool:
    """Evita poluir a aba de conferência da Movida com zeros sem valor.

    A medição da Movida traz muitas colunas auxiliares preenchidas com 0,00
    quando não houve cobrança. Para conferência manual, fica melhor manter essas
    células em branco. A coluna principal do rateio continua preservada.
    """
    if col_idx == col_valor:
        return False
    h = normalizar_texto(header)
    if h in {"fatura", "contrato", "pontos", "cod infracao", "empresa cnpj"}:
        return False
    termos_valor = (
        "valor",
        "taxa",
        "combustivel",
        "tanque",
        "km adicional",
        "avarias",
        "lavagem",
        "sem parar",
        "extras",
        "renegociacao",
        "desconto",
        "outras formas",
        "horas extras",
    )
    return any(termo in h for termo in termos_valor)


def _limpar_valores_base_movida(headers: list[Any], valores: list[Any], col_valor: int) -> list[Any]:
    limpos = list(valores)
    for idx, valor in enumerate(limpos, start=1):
        if not _deve_ocultar_zero_na_base(headers[idx - 1] if idx - 1 < len(headers) else "", idx, col_valor):
            continue
        if para_decimal(valor) == Decimal("0.00"):
            limpos[idx - 1] = None
    return limpos

def _identificador_fatura(linhas: list[dict[str, Any]], fallback: str) -> str:
    for linha in linhas:
        fatura = linha.get("FATURA")
        if fatura not in (None, ""):
            texto = str(fatura).strip()
            # preservar zero à esquerda em saídas do tipo 08321022
            digitos = re.sub(r"\D", "", texto)
            if digitos:
                return digitos.zfill(8) if len(digitos) <= 8 else digitos
    return fallback


def _processar_aba(wb, caminho_entrada: Path, pasta_saida: Path, tipo: str, overwrite: bool) -> RateioArquivo | None:
    ws = _aba_por_tipo(wb, tipo)
    if ws is None:
        return None

    config = CONFIG_ABAS[tipo]
    headers = _headers(ws)
    col_valor = _coluna(headers, config["coluna_valor"])
    col_cc = _coluna(headers, config["coluna_cc"])
    col_fatura = _coluna(headers, config["coluna_fatura"])
    col_nome = _coluna(headers, config["coluna_nome"])

    if not col_valor:
        raise ErroRateio(f"Movida/{config['subtipo']}: não encontrei a coluna de valor.")
    if not col_cc:
        raise ErroRateio(f"Movida/{config['subtipo']}: não encontrei a coluna CENTRO DE CUSTO.")

    linhas: list[LinhaRateio] = []
    dados_base: list[list[Any]] = []
    linhas_fatura: list[dict[str, Any]] = []
    total_arquivo = Decimal("0.00")
    valor_sem_cc = Decimal("0.00")
    linhas_sem_cc = 0
    linhas_ignoradas_total = 0

    for r in range(2, ws.max_row + 1):
        valores = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if _linha_total_ou_vazia(valores, col_valor):
            if any(v not in (None, "") for v in valores):
                linhas_ignoradas_total += 1
            continue

        valor = para_decimal(valores[col_valor - 1])
        cc_original = valores[col_cc - 1]
        cc = extrair_cc_de_texto(cc_original)
        descricao = descricao_por_cc_ou_texto(cc, cc_original)
        incluir = cc is not None and valor != Decimal("0.00")
        if not cc and valor != Decimal("0.00"):
            linhas_sem_cc += 1
            valor_sem_cc += decimal_moeda(valor)

        linha_dict = {str(headers[i] or f"COLUNA_{i+1}"): valores[i] for i in range(len(headers))}
        linhas_fatura.append(linha_dict)
        total_arquivo += decimal_moeda(valor)
        linhas.append(
            LinhaRateio(
                cc_original=cc_original,
                cc_ajustado=cc,
                cc_descricao=descricao,
                valor=valor,
                dados=linha_dict,
                incluir_no_rateio=incluir,
                motivo_ignorado=None if incluir else "SEM CC" if not cc else "VALOR ZERO",
            )
        )
        valores_base = _limpar_valores_base_movida(headers, valores, col_valor)
        dados_base.append(valores_base + [int(cc) if cc and cc.isdigit() else cc, descricao])

    if not linhas:
        return None

    identificador = _identificador_fatura(linhas_fatura, fallback="SEM_FATURA")
    nome_saida = f"RATEIO_{nome_seguro_rateio(identificador)}.xlsx"
    caminho_saida = pasta_saida / nome_saida
    if caminho_saida.exists() and not overwrite:
        base = caminho_saida.with_suffix("")
        contador = 2
        while caminho_saida.exists():
            caminho_saida = base.with_name(f"{base.name}_{contador}").with_suffix(".xlsx")
            contador += 1
        nome_saida = caminho_saida.name

    totais = somar_por_cc(linhas)
    criar_workbook_rateio(
        caminho_saida=caminho_saida,
        titulo_rateio=f"RATEIO {identificador}",
        nome_aba_base=config["subtipo"],
        headers_base=[str(h or "") for h in headers] + ["CC AJUSTADO", "CC DESCRIÇÃO"],
        linhas_base=dados_base,
        totais=totais,
    )

    avisos: list[str] = []
    if linhas_sem_cc:
        avisos.append(
            f"{linhas_sem_cc} linha(s) com valor ficaram sem CC para conferência manual "
            f"({formatar_moeda_br(valor_sem_cc)} fora do total rateado)."
        )
    if linhas_ignoradas_total:
        avisos.append(f"{linhas_ignoradas_total} linha(s) de total/rodapé foram ignoradas para não duplicar o valor.")

    total_rateado = sum(totais.values(), Decimal("0.00"))
    return RateioArquivo(
        fornecedor="Movida",
        subtipo=config["subtipo"],
        arquivo_origem=caminho_entrada.name,
        identificador=identificador,
        nome_saida=nome_saida,
        arquivo_saida=caminho_saida,
        total_rateado=decimal_moeda(total_rateado),
        total_arquivo=decimal_moeda(total_arquivo),
        valor_sem_cc=decimal_moeda(valor_sem_cc),
        linhas_processadas=len(linhas),
        linhas_rateadas=sum(1 for l in linhas if l.incluir_no_rateio and l.cc_ajustado),
        linhas_sem_cc=linhas_sem_cc,
        quantidade_cc=len(totais),
        avisos=avisos,
    )


def processar(caminho_entrada: Path, pasta_saida: Path, overwrite: bool = True) -> list[RateioArquivo]:
    caminho_entrada = Path(caminho_entrada)
    if caminho_entrada.suffix.lower() != ".xlsx":
        raise ErroRateio("Movida aceita apenas arquivos .xlsx de medição.")

    wb = load_workbook(caminho_entrada, data_only=True)
    resultados: list[RateioArquivo] = []
    try:
        for tipo in ["LOCACAO", "MULTAS"]:
            resultado = _processar_aba(wb, caminho_entrada, pasta_saida, tipo, overwrite)
            if resultado:
                resultados.append(resultado)
    finally:
        wb.close()

    if not resultados:
        raise ErroRateio("Não encontrei abas LOCAÇÃO ou MULTAS com dados para ratear.")
    return resultados
