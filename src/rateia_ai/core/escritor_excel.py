from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.rateia_ai.core.modelo import LinhaRateio
from src.rateia_ai.core.normalizacao import decimal_moeda
from src.rateia_ai.core.rateador import chave_ordenacao_cc

COR_VERDE = "007965"
COR_PRETO = "000000"
COR_BRANCO = "FFFFFF"
BORDA_FINA = Side(style="thin", color="A6A6A6")
BORDA = Border(left=BORDA_FINA, right=BORDA_FINA, top=BORDA_FINA, bottom=BORDA_FINA)


def _aplicar_estilo_rateio(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 15
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18


def criar_workbook_rateio(
    caminho_saida: Path,
    titulo_rateio: str,
    nome_aba_base: str,
    headers_base: list[str],
    linhas_base: list[list[Any]],
    totais: dict[str, Decimal],
) -> None:
    """Cria um Excel com aba RATEIO e uma aba de base conferível."""
    wb = Workbook()
    ws_rateio = wb.active
    ws_rateio.title = "RATEIO"
    _aplicar_estilo_rateio(ws_rateio)

    ws_rateio.merge_cells("B2:C2")
    ws_rateio["B2"] = titulo_rateio
    ws_rateio["B2"].font = Font(bold=True, size=12)
    ws_rateio["B2"].alignment = Alignment(horizontal="center")

    ws_rateio["B3"] = "CC AJUSTADO"
    ws_rateio["C3"] = "VALOR"
    for cel in (ws_rateio["B3"], ws_rateio["C3"]):
        cel.font = Font(bold=True, color=COR_BRANCO)
        cel.fill = PatternFill("solid", fgColor=COR_VERDE)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = BORDA

    linha = 4
    total_geral = Decimal("0.00")
    for cc in sorted(totais, key=chave_ordenacao_cc):
        valor = decimal_moeda(totais[cc])
        total_geral += valor
        ws_rateio.cell(linha, 2).value = int(cc) if str(cc).isdigit() else cc
        ws_rateio.cell(linha, 3).value = float(valor)
        ws_rateio.cell(linha, 2).number_format = "0"
        ws_rateio.cell(linha, 3).number_format = '#,##0.00'
        ws_rateio.cell(linha, 2).alignment = Alignment(horizontal="center")
        ws_rateio.cell(linha, 3).alignment = Alignment(horizontal="right")
        ws_rateio.cell(linha, 2).border = BORDA
        ws_rateio.cell(linha, 3).border = BORDA
        linha += 1

    ws_rateio.cell(linha, 2).value = "Total Geral"
    ws_rateio.cell(linha, 3).value = float(decimal_moeda(total_geral))
    for col in (2, 3):
        cell = ws_rateio.cell(linha, col)
        cell.font = Font(bold=True, color=COR_BRANCO)
        cell.fill = PatternFill("solid", fgColor=COR_PRETO)
        cell.border = BORDA
        cell.alignment = Alignment(horizontal="center" if col == 2 else "right")
    ws_rateio.cell(linha, 3).number_format = '"R$" #,##0.00;[Red]-"R$" #,##0.00'

    ws_base = wb.create_sheet(nome_aba_base[:31] or "BASE")
    ws_base.sheet_view.showGridLines = False
    ws_base.append(headers_base)
    for row in linhas_base:
        ws_base.append(row)

    header_fill = PatternFill("solid", fgColor=COR_VERDE)
    for cell in ws_base[1]:
        cell.font = Font(bold=True, color=COR_BRANCO)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA
    ws_base.freeze_panes = "A2"
    ws_base.auto_filter.ref = ws_base.dimensions

    for row in ws_base.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'

    # Larguras simples e seguras para leitura.
    for col_idx, header in enumerate(headers_base, start=1):
        texto = str(header or "")
        largura = min(max(len(texto) + 2, 10), 35)
        if texto.upper() in {"CC AJUSTADO", "CC DESCRIÇÃO"}:
            largura = 16
        ws_base.column_dimensions[get_column_letter(col_idx)].width = largura

    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho_saida)
