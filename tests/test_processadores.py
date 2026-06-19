from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.rateia_ai.adapters import biztrip, movida, uber
from src.rateia_ai.core.normalizacao import descricao_por_cc_ou_texto, extrair_cc_de_texto, para_decimal


EXEMPLOS = Path("/mnt/data/exemplos_work/exemplos")


def test_extrair_cc_de_texto():
    assert extrair_cc_de_texto("1.10622 - GE AEROSPACE - MRO") == "110622"
    assert extrair_cc_de_texto("valmet") == "110620"
    assert extrair_cc_de_texto("Presidência") == "510101"
    assert extrair_cc_de_texto("Comercial Corporativo") == "410102"


def test_descricao_cc_usa_tabela_oficial():
    assert descricao_por_cc_ou_texto("110612", "BRACELL antigo") == "KOMATSU"
    assert descricao_por_cc_ou_texto("110614", "KLABIN antigo") == "DAIICHI SANKYO"
    assert descricao_por_cc_ou_texto("110622", "GE AEROSPACE") == "GE MRO"
    assert descricao_por_cc_ou_texto("410102", "Comercial Corporativo") == "COMERCIAL CORPORATIVO"


def test_para_decimal():
    assert para_decimal("R$ 1.234,56") == Decimal("1234.56")
    assert para_decimal("1234.56") == Decimal("1234.56")


def test_movida_abril(tmp_path):
    origem = EXEMPLOS / "Medi#U00e7#U00e3o Rio Verde Engenharia -  ABRIL 2026.xlsx"
    if not origem.exists():
        return
    resultados = movida.processar(origem, tmp_path)
    por_subtipo = {r.subtipo: r for r in resultados}
    assert por_subtipo["LOCAÇÃO"].identificador == "08321022"
    assert por_subtipo["LOCAÇÃO"].total_rateado == Decimal("134698.06")
    assert por_subtipo["LOCAÇÃO"].valor_sem_cc == Decimal("3835.63")
    wb = load_workbook(tmp_path / "RATEIO_08321022.xlsx", data_only=True)
    ws = wb["LOCAÇÃO"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_cc = headers.index("CC AJUSTADO") + 1
    col_desc = headers.index("CC DESCRIÇÃO") + 1
    descricoes = {str(ws.cell(r, col_cc).value): ws.cell(r, col_desc).value for r in range(2, ws.max_row + 1)}
    assert descricoes["110612"] == "KOMATSU"
    assert descricoes["110614"] == "DAIICHI SANKYO"
    wb.close()
    assert por_subtipo["MULTAS"].identificador == "08321021"
    assert por_subtipo["MULTAS"].total_rateado == Decimal("944.24")


def test_movida_base_nao_preenche_zeros_auxiliares(tmp_path):
    origem = EXEMPLOS / "Medi#U00e7#U00e3o Rio Verde Engenharia -  ABRIL 2026.xlsx"
    if not origem.exists():
        return
    movida.processar(origem, tmp_path)
    wb = load_workbook(tmp_path / "RATEIO_08321022.xlsx", data_only=True)
    ws = wb["LOCAÇÃO"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_horas_extras = headers.index("HORAS EXTRAS DIARIA") + 1
    col_total_fatura = headers.index("TOTAL FATURA") + 1
    assert ws.cell(2, col_horas_extras).value is None
    assert ws.cell(2, col_total_fatura).value == 13.78
    wb.close()


def test_uber_maio(tmp_path):
    origem = EXEMPLOS / "b2dec9c9-32d7-5497-ae67-0ca5aa692978_u4b_transaction_csv (1).csv"
    if not origem.exists():
        return
    resultado = uber.processar(origem, tmp_path)[0]
    assert resultado.identificador == "B2DEC9C932"
    assert resultado.total_rateado == Decimal("28399.53")


def test_uber_datas_convertidas_para_padrao_brasileiro(tmp_path):
    origem = EXEMPLOS / "b2dec9c9-32d7-5497-ae67-0ca5aa692978_u4b_transaction_csv (1).csv"
    if not origem.exists():
        return
    uber.processar(origem, tmp_path)
    wb = load_workbook(tmp_path / "RATEIO_B2DEC9C932.xlsx", data_only=True)
    ws = wb["BASE"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_data_local = headers.index("Data da solicitação (local)") + 1
    col_timestamp = headers.index("Registro de data e hora da transação (UTC)") + 1
    assert ws.cell(2, col_data_local).value == "30/04/2026"
    assert ws.cell(2, col_timestamp).value == "01/05/2026 00:49:07"
    wb.close()


def test_biztrip_ft(tmp_path):
    origem = Path("/mnt/data/PLANILHA FT 39336.xlsx")
    if not origem.exists():
        return
    resultado = biztrip.processar(origem, tmp_path)[0]
    assert resultado.identificador == "39336"
    assert resultado.total_rateado == Decimal("87.38")
